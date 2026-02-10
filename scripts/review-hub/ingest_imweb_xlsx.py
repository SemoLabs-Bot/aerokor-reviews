#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import os
import re
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

# Ensure workspace root is on sys.path so `import review_hub` works when invoked as a script.
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from review_hub.sheets_client import GogSheetsClient


WORKSPACE_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))


def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def norm_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def parse_dt(s: str) -> str:
    s = s.strip()
    if not s:
        return ""
    # Input sample: 2026-02-10 10:51:49
    # Keep as-is.
    return s


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(description="Ingest Imweb review export XLSX into Review Hub main_review sheet.")
    ap.add_argument("xlsx", help="Path to Imweb XLSX (구매평_기본 양식_*.xlsx)")
    ap.add_argument("--brand", default="OLLY")
    ap.add_argument("--platform", default="imweb")
    ap.add_argument("--sheet-id", default=None)
    ap.add_argument("--account", default=None)
    ap.add_argument("--tab", default="main_review")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--max-rows", type=int, default=0, help="0=all")
    ap.add_argument("--chunk", type=int, default=200)
    ap.add_argument(
        "--filter-product-regex",
        default="",
        help="Optional regex to filter 상품명 (e.g., '^OLLY')",
    )

    args = ap.parse_args()

    # Load sink config for defaults
    sink_path = os.path.join(WORKSPACE_ROOT, "config", "review-hub", "google-sheets.sink.json")
    with open(sink_path, "r", encoding="utf-8") as f:
        sink = json.load(f)

    account = args.account or sink["account"]
    sheet_id = args.sheet_id or sink["sheetId"]

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb.active

    # Header row is row 1
    headers = [norm_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    def get(row: list[Any], name: str) -> str:
        i = idx.get(name)
        if i is None or i >= len(row):
            return ""
        return norm_str(row[i])

    product_re = re.compile(args.filter_product_regex) if args.filter_product_regex else None

    now = datetime.now().astimezone()
    collected_date = now.date().isoformat()
    collected_at = now.isoformat(timespec="seconds")

    out_rows: list[list[object]] = []

    for n, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if args.max_rows and n > args.max_rows:
            break

        row = list(r)
        product_name = get(row, "상품명")
        if product_re and (not product_re.search(product_name)):
            continue

        review_id = get(row, "글번호")
        body = get(row, "글 내용")
        author = get(row, "작성자")
        review_date = parse_dt(get(row, "작성시각"))
        rating = get(row, "평점")

        # Skip empty bodies / empty ids
        if not review_id and not body:
            continue

        body_hash = sha256_hex(body)

        # dedup_key per sink config: composite_sha256(fields=[platform, product_url, author, review_date, body_hash])
        product_url = ""
        dedup_material = "|".join([
            args.platform,
            product_url,
            author,
            review_date,
            body_hash,
        ])
        dedup_key = sha256_hex(dedup_material)

        # Our canonical columns (A:O):
        # collected_date, collected_at, brand, platform, product_name, product_url,
        # review_id, review_date, rating, author, title, body, body_hash, dedup_key, source_url
        out_rows.append(
            [
                collected_date,
                collected_at,
                args.brand,
                args.platform,
                product_name,
                product_url,
                review_id,
                review_date,
                rating,
                author,
                "",  # title
                body,
                body_hash,
                dedup_key,
                "",  # source_url
            ]
        )

    # Write audit artifact
    out_dir = os.path.join(WORKSPACE_ROOT, "outputs", "review-hub")
    os.makedirs(out_dir, exist_ok=True)
    audit_path = os.path.join(out_dir, f"imweb_xlsx_extract_{args.brand}_{collected_date}.json")
    with open(audit_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "source_xlsx": os.path.abspath(args.xlsx),
                "generated_at": collected_at,
                "brand": args.brand,
                "platform": args.platform,
                "count": len(out_rows),
                "columns": sink.get("columns"),
                "rows": out_rows[:50],
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    if args.dry_run:
        print(json.dumps({"dry_run": True, "count": len(out_rows), "audit": audit_path}, ensure_ascii=False))
        return

    client = GogSheetsClient(account=account, spreadsheet_id=sheet_id)

    # Append in chunks using fixed append to avoid table heuristic.
    start_row = 3
    start_col = "A"
    end_col = "O"

    written_ranges: list[str] = []
    for i in range(0, len(out_rows), args.chunk):
        chunk = out_rows[i : i + args.chunk]
        rng = client.append_fixed(
            tab=args.tab,
            start_row=start_row,
            start_col=start_col,
            end_col=end_col,
            values_2d=chunk,
            sentinel_col="A",
            sentinel_regex=r"^\d{4}-\d{2}-\d{2}$",
            scan_max_rows=10000,
        )
        written_ranges.append(rng)

    print(
        json.dumps(
            {
                "ok": True,
                "count": len(out_rows),
                "audit": audit_path,
                "written_ranges": written_ranges[:3] + (["..."] if len(written_ranges) > 3 else []),
                "chunks": len(written_ranges),
                "sheet": sheet_id,
                "tab": args.tab,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
