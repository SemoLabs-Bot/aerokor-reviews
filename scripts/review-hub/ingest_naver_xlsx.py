#!/usr/bin/env python3
from __future__ import annotations

"""Ingest Naver review XLSX export into Review Hub main_review.

Input: ~/Downloads/네이버_리뷰.xlsx (columns like 상품명, 리뷰상세내용, 등록자, 리뷰등록일, 구매자평점, 리뷰글번호, 포토/영상)

Mapping -> main_review columns (A:O):
- collected_date, collected_at: now
- brand: inferred from 상품명 (default rules) or --brand override
- platform: naver
- product_name: 상품명
- product_url: (not provided) empty
- review_id: 리뷰글번호
- review_date: 리뷰등록일 (kept as-is)
- rating: 구매자평점
- author: 등록자
- title: empty (naver export doesn't provide distinct title)
- body: 리뷰상세내용
- body_hash: sha256(body)
- dedup_key: composite sha256(platform|product_url|author|review_date|body_hash) (same as sink)
- source_url: 포토/영상 URL if present (first URL)

Note: This script appends (does not dedupe). Run dedupe scripts after ingestion if needed.
"""

import hashlib
import json
import os
import re
import sys
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

WORKSPACE_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, WORKSPACE_ROOT)

from review_hub.sheets_client import GogSheetsClient


def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def norm(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def infer_brand(product_name: str) -> str:
    s = product_name.strip()
    if not s:
        return ""
    # Basic inference (extend as needed)
    if "올리" in s or s.upper().startswith("OLLY"):
        return "OLLY"
    return ""


def first_url(s: str) -> str:
    m = re.search(r"https?://\S+", s)
    return m.group(0) if m else ""


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Path to 네이버_리뷰.xlsx")
    ap.add_argument("--brand", default="", help="Override brand")
    ap.add_argument("--platform", default="naver")
    ap.add_argument("--sheet-id", default=None)
    ap.add_argument("--account", default=None)
    ap.add_argument("--tab", default="main_review")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--chunk", type=int, default=500)

    args = ap.parse_args()

    sink_path = os.path.join(WORKSPACE_ROOT, "config", "review-hub", "google-sheets.sink.json")
    sink = json.load(open(sink_path, "r", encoding="utf-8"))
    account = args.account or sink["account"]
    sheet_id = args.sheet_id or sink["sheetId"]

    # read_only mode gives unreliable max_row/max_col for this export; load normally (file is small).
    wb = load_workbook(args.xlsx, read_only=False, data_only=True)
    ws = wb.active

    max_col = ws.max_column or 25
    max_row = ws.max_row or 1000
    headers = [norm(v) for v in next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True))]
    idx = {h: i for i, h in enumerate(headers) if h}

    def get(row: list[Any], name: str) -> str:
        i = idx.get(name)
        if i is None or i >= len(row):
            return ""
        return norm(row[i])

    now = datetime.now().astimezone()
    collected_date = now.date().isoformat()
    collected_at = now.isoformat(timespec="seconds")

    out_rows: list[list[object]] = []

    for r in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        row = list(r)
        product_name = get(row, "상품명")
        body = get(row, "리뷰상세내용")
        author = get(row, "등록자")
        review_date = get(row, "리뷰등록일")
        rating = get(row, "구매자평점")
        review_id = get(row, "리뷰글번호")
        media = get(row, "포토/영상")

        if not (product_name or body or review_id):
            continue

        brand = args.brand or infer_brand(product_name)
        body_hash = sha256_hex(body)
        product_url = ""
        dedup_material = "|".join([args.platform, product_url, author, review_date, body_hash])
        dedup_key = sha256_hex(dedup_material)
        source_url = first_url(media)

        out_rows.append(
            [
                collected_date,
                collected_at,
                brand,
                args.platform,
                product_name,
                product_url,
                review_id,
                review_date,
                rating,
                author,
                "",  # title
                body,
                body_hash,
                dedup_key,
                source_url,
            ]
        )

    out_dir = os.path.join(WORKSPACE_ROOT, "outputs", "review-hub")
    os.makedirs(out_dir, exist_ok=True)
    audit_path = os.path.join(out_dir, f"naver_xlsx_extract_{collected_date}.json")
    json.dump(
        {
            "source_xlsx": os.path.abspath(args.xlsx),
            "generated_at": collected_at,
            "count": len(out_rows),
            "rows_preview": out_rows[:20],
        },
        open(audit_path, "w", encoding="utf-8"),
        ensure_ascii=False,
        indent=2,
    )

    if args.dry_run:
        print(json.dumps({"dry_run": True, "count": len(out_rows), "audit": audit_path}, ensure_ascii=False))
        return

    client = GogSheetsClient(account=account, spreadsheet_id=sheet_id)

    # Compute next row once
    start_row = 3
    colA = client.get(f"{args.tab}!A{start_row}:A200000")
    last = start_row - 1
    pat = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    for i, rr in enumerate(colA, start=start_row):
        v = rr[0] if rr else ""
        if isinstance(v, str) and pat.search(v.strip()):
            last = i
    next_row = last + 1

    written = []
    for i in range(0, len(out_rows), args.chunk):
        chunk = out_rows[i : i + args.chunk]
        s = next_row
        e = s + len(chunk) - 1
        rng = f"{args.tab}!A{s}:O{e}"
        client.update(rng, chunk)
        written.append(rng)
        next_row = e + 1

    print(json.dumps({"ok": True, "count": len(out_rows), "audit": audit_path, "chunks": len(written), "written_ranges": written[:3] + (["..."] if len(written) > 3 else [])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
