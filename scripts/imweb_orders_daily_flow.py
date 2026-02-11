#!/usr/bin/env python3
"""Imweb orders XLSX -> Clean -> Pivot -> Google Sheets.

Usage examples:
  python3 scripts/imweb_orders_daily_flow.py --input ~/Downloads/기본_양식_20260210163212.xlsx
  python3 scripts/imweb_orders_daily_flow.py --from-downloads --since-minutes 30

This script:
- Loads the XLSX (Imweb export '기본 양식')
- Cleans rows:
  - drop rows where 취소사유 not empty
  - drop rows where 반품사유 not empty
  - drop rows where 주문상태 == 결제대기 (best-effort automation)
- Creates pivot:
  Rows: 주문일 -> 상품명 -> 옵션명
  Values: 구매수량 SUM, 품목실결제가 SUM
  Adds:
    - 실판매수량 (옵션명 contains '1+1' => 구매수량*2)
    - 객단가 = 품목실결제가 / 실판매수량
- Appends cleaned rows to Google Sheet tab A
- Clears & rewrites pivot tab B

Requires:
- gog CLI authenticated (Sheets scope)
- SMTP configured for email notifications (same style as meta flow)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None


DEFAULT_SHEET_ID = "1eSNPcGuOEdj4Lj-2tCQcdcFR-C1Z7sIS2ULbhezGtTo"
DEFAULT_TAB_RAW = "올리_아임웹_주문원본"
DEFAULT_TAB_PIVOT = "올리_아임웹_피벗집계"

DEFAULT_EMAIL_TO = "realtiger@wekeepgrowing.com"
DEFAULT_EMAIL_FROM = os.environ.get("IMWEB_EMAIL_FROM") or os.environ.get("GOG_ACCOUNT") or "semolabsbot@gmail.com"


@dataclass
class RunStats:
    input_path: str
    rows_total: int
    rows_cancel_dropped: int
    rows_return_dropped: int
    rows_pending_dropped: int
    rows_after_clean: int
    pivot_rows: int


def sh(cmd: list[str], *, text=True) -> str:
    p = subprocess.run(cmd, text=text, capture_output=True)
    if p.returncode != 0:
        raise RuntimeError(f"command failed: {' '.join(cmd)}\nstdout: {p.stdout}\nstderr: {p.stderr}")
    return p.stdout


def pick_latest_xlsx(downloads_dir: Path, since_minutes: int, prefix: str = "기본_양식_") -> Path:
    """Pick latest Imweb export XLSX from Downloads.

    NOTE: macOS can produce filenames in decomposed Unicode (NFD), e.g.
    '기본_양식_' instead of '기본_양식_'. We normalize to NFC for matching.
    """

    import unicodedata

    since_ts = datetime.now().timestamp() - since_minutes * 60
    cand: list[tuple[float, Path]] = []

    # Accept both 'prefix' and its NFD variant
    prefix_nfc = unicodedata.normalize("NFC", prefix)
    prefix_nfd = unicodedata.normalize("NFD", prefix)

    for p in downloads_dir.glob("*.xlsx"):
        try:
            st = p.stat()
        except FileNotFoundError:
            continue
        if st.st_mtime < since_ts:
            continue

        name_nfc = unicodedata.normalize("NFC", p.name)
        name_nfd = unicodedata.normalize("NFD", p.name)

        if name_nfc.startswith(prefix_nfc) or name_nfd.startswith(prefix_nfd):
            cand.append((st.st_mtime, p))

    if not cand:
        raise FileNotFoundError(
            f"No recent XLSX found in {downloads_dir} within last {since_minutes} minutes matching {prefix}*.xlsx"
        )

    cand.sort(reverse=True)
    return cand[0][1]


def normalize_str(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() in {"nan", "none"} else s


def to_number(x) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = normalize_str(x)
    if not s:
        return 0.0
    # remove commas
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def load_xlsx_rows(path: Path) -> tuple[list[str], list[dict]]:
    """Load rows from an Imweb export workbook.

    The downloaded '기본_양식' files sometimes contain:
    - a raw data sheet (wide table)
    - a pivot/summary sheet

    We must pick the sheet that actually contains the raw export columns.
    Heuristic: choose the first sheet whose header row contains core fields.
    """

    wb = load_workbook(path, data_only=True, read_only=True)

    required_core = {"주문일", "상품명", "옵션명", "구매수량", "품목실결제가"}

    picked = None
    picked_headers: list[str] | None = None

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        headers = [normalize_str(h) for h in header_row]
        header_set = {h for h in headers if h}

        if required_core.issubset(header_set):
            picked = ws
            picked_headers = headers
            break

    if picked is None:
        # fallback: first sheet
        picked = wb.worksheets[0]
        rows_iter = picked.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Empty xlsx")
        picked_headers = [normalize_str(h) for h in header_row]

    headers = picked_headers

    data: list[dict] = []
    for row in picked.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(v is None or normalize_str(v) == "" for v in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            if i < len(row):
                d[h] = row[i]
        data.append(d)

    return headers, data


def clean_orders(rows: list[dict]) -> tuple[list[dict], dict]:
    total = len(rows)

    def is_nonempty(r: dict, col: str) -> bool:
        return normalize_str(r.get(col)) != ""

    cancel = 0
    ret = 0
    pending = 0
    out: list[dict] = []

    for r in rows:
        drop = False
        if is_nonempty(r, "취소사유"):
            cancel += 1
            drop = True
        if is_nonempty(r, "반품사유"):
            ret += 1
            drop = True
        if normalize_str(r.get("주문상태")) == "결제대기":
            pending += 1
            drop = True
        if not drop:
            out.append(r)

    stats = {
        "rows_total": total,
        "rows_cancel_dropped": cancel,
        "rows_return_dropped": ret,
        "rows_pending_dropped": pending,
        "rows_after_clean": len(out),
    }
    return out, stats


def is_one_plus_one_set(option_name: str) -> bool:
    """Return True if this option should be treated as a 1+1 set (sale qty *2).

    IMPORTANT: Template pivot (기본_양식 Sheet2) does NOT double-count every option
    that contains '1+1'. It appears to double only explicit '1+1SET' bundles.

    Examples that should double:
    - '구성선택 : [1+1SET] OLSB24W + OLSB24W'

    Examples that should NOT double:
    - '구성선택 : [기본구성] 1+1 구성'
    """

    s = normalize_str(option_name)
    s_norm = s.replace(" ", "")
    return "1+1SET" in s_norm or "[1+1SET]" in s_norm


def build_pivot_template(rows: list[dict], *, include_order_date: bool) -> list[list]:
    """Build pivot rows matching the Excel template look.

    Output columns (A:E):
      행 레이블 | 합계 : 구매수량 | 합계 : 품목실결제가 | 객단가 | 실판매수량

    Hierarchy:
      - If include_order_date: 주문일 -> 상품명 -> 옵션명
      - Else: 상품명 -> 옵션명

    Notes:
    - Option rows do NOT show 객단가/실판매수량 in template.
    - Product rows DO show 객단가 + 실판매수량 (with 1+1SET doubling rule).
    - We emulate indentation by prefixing spaces.
    """

    required = ["주문일", "상품명", "옵션명", "구매수량", "품목실결제가"]
    for c in required:
        if not any(c in r for r in rows):
            raise KeyError(f"Missing required column for pivot: {c}")

    from collections import defaultdict

    # Grouping structure
    # date_key -> product_key -> option_key -> sums
    def new_bucket():
        return {
            "qty": 0.0,
            "pay": 0.0,
            "sale_qty": 0.0,
            "opts": defaultdict(lambda: {"qty": 0.0, "pay": 0.0, "sale_qty": 0.0}),
        }

    by_date = defaultdict(lambda: defaultdict(new_bucket))

    for r in rows:
        od = normalize_str(r.get("주문일"))[:10]
        prod = normalize_str(r.get("상품명"))
        opt = normalize_str(r.get("옵션명")) or "(비어 있음)"
        qty = to_number(r.get("구매수량"))
        pay = to_number(r.get("품목실결제가"))
        sale_qty = qty * (2.0 if is_one_plus_one_set(opt) else 1.0)

        date_key = od if include_order_date else "__SINGLE__"
        b = by_date[date_key][prod]
        b["qty"] += qty
        b["pay"] += pay
        b["sale_qty"] += sale_qty
        o = b["opts"][opt]
        o["qty"] += qty
        o["pay"] += pay
        o["sale_qty"] += sale_qty

    def aov(pay: float, sale_qty: float) -> float:
        return (pay / sale_qty) if sale_qty else 0.0

    out: list[list] = []
    # Template has two blank rows
    out.append(["", "", "", "", ""])
    out.append(["", "", "", "", ""])
    # Template header shows only first 3 labels; cols D/E are blank
    out.append(["행 레이블", "합계 : 구매수량", "합계 : 품목실결제가", "", ""])

    # Deterministic ordering
    date_keys = sorted(by_date.keys())

    grand_qty = grand_pay = grand_sale = 0.0

    for dk in date_keys:
        products = by_date[dk]

        # date-level subtotal if we include order date
        if include_order_date:
            date_qty = sum(products[p]["qty"] for p in products)
            date_pay = sum(products[p]["pay"] for p in products)
            date_sale = sum(products[p]["sale_qty"] for p in products)
            # Keep date row similar style (no indentation)
            out.append([dk, date_qty, date_pay, aov(date_pay, date_sale), date_sale])

        for prod in sorted(products.keys()):
            b = products[prod]
            grand_qty += b["qty"]
            grand_pay += b["pay"]
            grand_sale += b["sale_qty"]

            opts_sorted = sorted(b["opts"].keys())
            only_empty_opt = (len(opts_sorted) == 1 and opts_sorted[0] == "(비어 있음)")

            # Template quirk: if the only option is (비어 있음), the product row shows blank aov/sale,
            # and the (비어 있음) row shows aov/sale instead.
            prod_aov = "" if only_empty_opt else aov(b["pay"], b["sale_qty"])
            prod_sale = "" if only_empty_opt else b["sale_qty"]

            out.append([prod, b["qty"], b["pay"], prod_aov, prod_sale])

            for opt in opts_sorted:
                o = b["opts"][opt]
                if only_empty_opt:
                    out.append([opt, o["qty"], o["pay"], aov(b["pay"], b["sale_qty"]), b["sale_qty"]])
                else:
                    out.append([opt, o["qty"], o["pay"], "", ""])

    out.append(["총합계", grand_qty, grand_pay, "", ""])
    return out


def sheets_get_header(sheet_id: str, tab: str, account: str | None) -> list[str] | None:
    args = ["gog", "sheets", "get", sheet_id, f"{tab}!A1:ZZ1", "--json", "--no-input"]
    if account:
        args += ["--account", account]
    raw = sh(args)
    j = json.loads(raw)
    values = j.get("values") or []
    if not values:
        return None
    return [normalize_str(x) for x in values[0] if normalize_str(x) != ""]


def sheets_append(sheet_id: str, tab: str, rows: list[list], account: str | None):
    # Append to a wide range; Google ignores extra columns beyond provided.
    args = [
        "gog",
        "sheets",
        "append",
        sheet_id,
        f"{tab}!A:ZZ",
        "--values-json",
        json.dumps(rows, ensure_ascii=False),
        "--insert",
        "INSERT_ROWS",
        "--no-input",
    ]
    if account:
        args += ["--account", account]
    sh(args)


def sheets_clear(sheet_id: str, tab: str, account: str | None):
    args = ["gog", "sheets", "clear", sheet_id, f"{tab}!A:ZZ", "--no-input"]
    if account:
        args += ["--account", account]
    sh(args)


def sheets_update(sheet_id: str, tab: str, start_cell: str, rows: list[list], account: str | None):
    args = [
        "gog",
        "sheets",
        "update",
        sheet_id,
        f"{tab}!{start_cell}",
        "--values-json",
        json.dumps(rows, ensure_ascii=False),
        "--input",
        "USER_ENTERED",
        "--no-input",
    ]
    if account:
        args += ["--account", account]
    sh(args)


def a1_col(n: int) -> str:
    """1-indexed column number -> A1 column letters."""

    if n < 1:
        raise ValueError("column must be >= 1")
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def sheets_format(sheet_id: str, tab: str, a1_range: str, cell_format: dict, fields: str, account: str | None):
    args = [
        "gog",
        "sheets",
        "format",
        sheet_id,
        f"{tab}!{a1_range}",
        "--format-json",
        json.dumps(cell_format, ensure_ascii=False),
        "--format-fields",
        fields,
        "--no-input",
    ]
    if account:
        args += ["--account", account]
    sh(args)


def send_email(*, sheet_id: str, label: str, email_to: str, email_from: str) -> str:
    """Send completion email (same address as meta flow by default)."""

    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
    subject = f"[아임웹 주문 리포트] {label} 업데이트 완료 (Google Sheets 링크)"
    body = (
        "안녕하세요.\n"
        f"아임웹 주문 리포트({label}, KST) 기준으로 데이터 업데이트 완료된 구글 시트 링크 공유드립니다.\n\n"
        f"- 문서 링크: {sheet_url}\n\n"
        "확인 부탁드립니다.\n"
        "감사합니다.\n"
        "semolabsbot\n"
    )

    out_dir = Path("out") / "imweb-orders-run" / datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "email_subject.txt").write_text(subject, encoding="utf-8")
    (out_dir / "email_body.txt").write_text(body, encoding="utf-8")

    # Reuse the same SMTP helper script used by meta flow.
    out = sh(
        [
            "node",
            "skills/imap-smtp-email/scripts/smtp.js",
            "send",
            "--to",
            email_to,
            "--subject",
            subject,
            "--body-file",
            str(out_dir / "email_body.txt"),
            "--from",
            email_from,
        ]
    )

    try:
        j = json.loads(out)
        return j.get("messageId", "")
    except Exception:
        return out.strip()[:500]


def chunked(seq: list[list], n: int):
    for i in range(0, len(seq), n):
        yield seq[i : i + n]


def excel_serial_to_datetime(x: float):
    """Convert Excel serial date (days since 1899-12-30) to datetime.

    Google Sheets uses the same epoch, so formatting the cell as DATE/DATETIME
    would also work, but we normalize to a readable string to avoid ambiguity.
    """

    from datetime import datetime, timedelta

    base = datetime(1899, 12, 30)
    return base + timedelta(days=float(x))


def rows_to_matrix(rows: list[dict], cols: list[str]) -> list[list]:
    out: list[list] = []
    for r in rows:
        row_out = []
        for c in cols:
            v = r.get(c, "")
            if c == "주문일":
                try:
                    # openpyxl may give datetime or Excel serial float.
                    if hasattr(v, "strftime"):
                        v = v.strftime("%Y-%m-%d %H:%M")
                    else:
                        s = normalize_str(v)
                        if s and re.fullmatch(r"\d+(\.\d+)?", s):
                            dt = excel_serial_to_datetime(float(s))
                            v = dt.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    pass
            row_out.append(v)
        out.append(row_out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", help="Path to imweb xlsx export")
    ap.add_argument("--from-downloads", action="store_true", help="Pick latest 기본_양식_*.xlsx from ~/Downloads")
    ap.add_argument("--downloads-dir", default=str(Path.home() / "Downloads"))
    ap.add_argument("--since-minutes", type=int, default=60)

    ap.add_argument("--sheet-id", default=DEFAULT_SHEET_ID)
    ap.add_argument("--tab-raw", default=DEFAULT_TAB_RAW)
    ap.add_argument("--tab-pivot", default=DEFAULT_TAB_PIVOT)
    ap.add_argument("--account", default=os.environ.get("GOG_ACCOUNT"))

    ap.add_argument("--email", action="store_true", help="Send completion email after updating Sheets")
    ap.add_argument("--email-to", default=DEFAULT_EMAIL_TO)
    ap.add_argument("--email-from", default=DEFAULT_EMAIL_FROM)
    ap.add_argument("--email-label", default="", help="Optional label override in email subject")

    ap.add_argument("--pivot-mode", choices=["replace"], default="replace")
    ap.add_argument("--append-chunk", type=int, default=300)

    # Safety/testing
    ap.add_argument("--dry-run", action="store_true", help="Do not write to Google Sheets. Print pivot preview only.")
    ap.add_argument(
        "--pivot-include-order-date",
        choices=["auto", "yes", "no"],
        default="auto",
        help="Include 주문일 group in pivot. auto=include only when multiple distinct dates are present.",
    )

    args = ap.parse_args()

    if not args.input and not args.from_downloads:
        ap.error("Either --input or --from-downloads must be provided")

    if args.from_downloads:
        input_path = pick_latest_xlsx(Path(args.downloads_dir), args.since_minutes)
    else:
        input_path = Path(args.input).expanduser()

    headers, rows_raw = load_xlsx_rows(input_path)
    rows_clean, s = clean_orders(rows_raw)

    # Decide whether to include 주문일 group in pivot
    distinct_dates = sorted({normalize_str(r.get("주문일"))[:10] for r in rows_clean if normalize_str(r.get("주문일"))})
    include_date = False
    if args.pivot_include_order_date == "yes":
        include_date = True
    elif args.pivot_include_order_date == "no":
        include_date = False
    else:  # auto
        include_date = len(distinct_dates) > 1

    pivot_matrix = build_pivot_template(rows_clean, include_order_date=include_date)

    # DRY RUN: do not write to Sheets
    if args.dry_run:
        print(
            json.dumps(
                {
                    "status": "dry_run",
                    "input": str(input_path),
                    "rows_total": s["rows_total"],
                    "rows_after_clean": s["rows_after_clean"],
                    "dropped": {
                        "cancel": s["rows_cancel_dropped"],
                        "return": s["rows_return_dropped"],
                        "pending": s["rows_pending_dropped"],
                    },
                    "pivot_include_order_date": include_date,
                    "distinct_dates": distinct_dates,
                    "pivot_preview": pivot_matrix[:30],
                    "note": "DRY RUN ONLY: did not write to Google Sheets",
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    # Append cleaned rows to RAW tab
    sheet_header = sheets_get_header(args.sheet_id, args.tab_raw, args.account)
    if sheet_header:
        cols = [c for c in sheet_header if c in headers]
        if not cols:
            cols = [c for c in headers if c]
    else:
        cols = [c for c in headers if c]
        sheets_update(args.sheet_id, args.tab_raw, "A1", [cols], args.account)

    matrix = rows_to_matrix(rows_clean, cols)
    for chunk in chunked(matrix, args.append_chunk):
        sheets_append(args.sheet_id, args.tab_raw, chunk, args.account)

    # Apply header formatting
    # - RAW header: #f5f5a1
    # - Pivot header: #d4dde9
    raw_header_fmt = {
        "backgroundColor": {"red": 245 / 255, "green": 245 / 255, "blue": 161 / 255},
        "textFormat": {"bold": True},
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
    }
    pivot_header_fmt = {
        "backgroundColor": {"red": 212 / 255, "green": 221 / 255, "blue": 233 / 255},
        "textFormat": {"bold": True},
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
    }
    header_fields = "backgroundColor,textFormat.bold,horizontalAlignment,verticalAlignment,wrapStrategy"
    try:
        raw_last_col = a1_col(len(cols) if cols else 1)
        sheets_format(args.sheet_id, args.tab_raw, f"A1:{raw_last_col}1", raw_header_fmt, header_fields, args.account)

        # Ensure the 주문일 column displays as datetime even if legacy rows are stored as Excel serial numbers.
        # (We also normalize new appends to string, but old rows can remain numeric.)
        sheets_format(
            args.sheet_id,
            args.tab_raw,
            "AK2:AK20000",
            {"numberFormat": {"type": "DATE_TIME", "pattern": "yyyy-mm-dd hh:mm"}},
            "numberFormat",
            args.account,
        )
    except Exception:
        # Formatting should never block data writes.
        pass

    # Replace pivot tab (template-like layout)
    if args.pivot_mode == "replace":
        sheets_clear(args.sheet_id, args.tab_pivot, args.account)
        row_cursor = 1
        for chunk in chunked(pivot_matrix, 500):
            sheets_update(args.sheet_id, args.tab_pivot, f"A{row_cursor}", chunk, args.account)
            row_cursor += len(chunk)

        # Pivot header row is always at row 3 in our template.
        # User preference: header background should NOT cover the extra computed columns (D/E).
        try:
            sheets_format(args.sheet_id, args.tab_pivot, "A3:C3", pivot_header_fmt, header_fields, args.account)
            # Explicitly clear header background on computed columns area (D/E).
            sheets_format(
                args.sheet_id,
                args.tab_pivot,
                "D3:E3",
                {"backgroundColor": {"red": 1, "green": 1, "blue": 1}},
                "backgroundColor",
                args.account,
            )
        except Exception:
            pass

        # User preference: apply background color for grand total / subtotals.
        total_fmt = {
            "backgroundColor": {"red": 212 / 255, "green": 221 / 255, "blue": 233 / 255},
            "textFormat": {"bold": True},
        }
        total_fields = "backgroundColor,textFormat.bold"
        try:
            for i, row in enumerate(pivot_matrix, start=1):
                label = normalize_str(row[0] if row else "")
                if not label:
                    continue
                if label == "총합계" or "소계" in label:
                    # User preference: total/subtotal background should NOT cover computed columns (D/E)
                    sheets_format(args.sheet_id, args.tab_pivot, f"A{i}:C{i}", total_fmt, total_fields, args.account)
                    # Clear background on computed columns
                    sheets_format(
                        args.sheet_id,
                        args.tab_pivot,
                        f"D{i}:E{i}",
                        {"backgroundColor": {"red": 1, "green": 1, "blue": 1}},
                        "backgroundColor",
                        args.account,
                    )
        except Exception:
            pass

    stats = RunStats(
        input_path=str(input_path),
        rows_total=s["rows_total"],
        rows_cancel_dropped=s["rows_cancel_dropped"],
        rows_return_dropped=s["rows_return_dropped"],
        rows_pending_dropped=s["rows_pending_dropped"],
        rows_after_clean=s["rows_after_clean"],
        pivot_rows=int(max(0, len(pivot_matrix) - 3)),
    )

    log_dir = Path("logs/imweb-orders")
    log_dir.mkdir(parents=True, exist_ok=True)
    kst = ZoneInfo("Asia/Seoul") if ZoneInfo else timezone(timedelta(hours=9))
    ts_kst = datetime.now(tz=kst).strftime("%Y-%m-%d")
    log_path = log_dir / f"imweb-orders-{ts_kst}.json"
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(stats.__dict__, f, ensure_ascii=False, indent=2)

    email_message_id = ""
    if args.email:
        # Default label: single date or range if multiple dates.
        label = args.email_label.strip()
        if not label:
            if distinct_dates:
                label = distinct_dates[0] if len(distinct_dates) == 1 else f"{distinct_dates[0]}~{distinct_dates[-1]}"
            else:
                label = ts_kst
        email_message_id = send_email(
            sheet_id=args.sheet_id,
            label=label,
            email_to=args.email_to,
            email_from=args.email_from,
        )

    print(
        json.dumps(
            {
                "status": "ok",
                "input": stats.input_path,
                "rows_total": stats.rows_total,
                "rows_after_clean": stats.rows_after_clean,
                "dropped": {
                    "cancel": stats.rows_cancel_dropped,
                    "return": stats.rows_return_dropped,
                    "pending": stats.rows_pending_dropped,
                },
                "pivot_rows": stats.pivot_rows,
                "pivot_include_order_date": include_date,
                "distinct_dates": distinct_dates,
                "sheet": args.sheet_id,
                "tab_raw": args.tab_raw,
                "tab_pivot": args.tab_pivot,
                "email_message_id": email_message_id,
                "log": str(log_path),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(json.dumps({"status": "error", "error": str(e)}, ensure_ascii=False))
        raise
