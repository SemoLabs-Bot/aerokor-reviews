#!/usr/bin/env python3
"""Meta Ads → Google Sheets → Email daily flow.

Key rules
- Compute yesterday in Asia/Seoul (KST) as REPORT_DATE.
- Input XLSX must be available as a fixed path in Downloads:
    ~/Downloads/meta_ads_daily_export.xlsx
  For this run, if it's missing, we auto-copy the newest matching user download
  (e.g. ~/Downloads/일자별-보고서*.xlsx) into that fixed path.
- ALWAYS validate that the XLSX '일' column contains only REPORT_DATE.
  If mismatch: STOP (no Sheets update, no email).
- Append (not replace) into Google Sheets tab '메타보고서' with dedupe-by-date.
- Cleanup: delete the fixed Downloads file at the end (success or failure).

This script intentionally keeps an audit copy under out/meta-ads-run/.
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from zoneinfo import ZoneInfo

SHEET_ID = "1eSNPcGuOEdj4Lj-2tCQcdcFR-C1Z7sIS2ULbhezGtTo"
TAB = "메타보고서"
ACCOUNT = os.environ.get("GOG_ACCOUNT", "semolabsbot@gmail.com")

DOWNLOADS = Path.home() / "Downloads"
FIXED_XLSX = DOWNLOADS / "meta_ads_daily_export.xlsx"

OUT_DIR = Path("out") / "meta-ads-run" / datetime.now().strftime("%Y%m%d-%H%M%S")
OUT_DIR.mkdir(parents=True, exist_ok=True)

META_URL = (
    "https://adsmanager.facebook.com/adsmanager/reporting/business_view"
    "?act=1027479588305818&ads_manager_write_regions=true&business_id=200761308489380"
    "&selected_report_id=1149377253627776"
)


class FlowStop(Exception):
    pass


@dataclass
class ExtractResult:
    report_date: str
    values: list[list]


def kst_yesterday() -> str:
    now_kst = datetime.now(ZoneInfo("Asia/Seoul"))
    y = (now_kst - timedelta(days=1)).date()
    return y.isoformat()


def sh(cmd: list[str], *, text=True) -> str:
    return subprocess.check_output(cmd, text=text)


def gog_get(a1_range: str) -> dict:
    out = sh([
        "gog",
        "sheets",
        "get",
        SHEET_ID,
        a1_range,
        "--account",
        ACCOUNT,
        "--json",
        "--no-input",
    ])
    return json.loads(out)


def gog_update(a1_range: str, values: list[list], *, user_entered=True) -> None:
    values_json = json.dumps(values, ensure_ascii=False)
    cmd = [
        "gog",
        "sheets",
        "update",
        SHEET_ID,
        a1_range,
        "--values-json",
        values_json,
        "--account",
        ACCOUNT,
        "--no-input",
    ]
    if user_entered:
        cmd += ["--input", "USER_ENTERED"]
    subprocess.check_call(cmd)


def gog_format(a1_range: str, fmt_json: dict, fmt_fields: str) -> None:
    cmd = [
        "gog",
        "sheets",
        "format",
        SHEET_ID,
        a1_range,
        "--format-json",
        json.dumps(fmt_json),
        "--format-fields",
        fmt_fields,
        "--account",
        ACCOUNT,
        "--no-input",
    ]
    subprocess.check_call(cmd)


def find_newest_user_download() -> Path | None:
    # Most common pattern observed: "일자별-보고서 (n).xlsx"
    candidates = list(DOWNLOADS.glob("일자별-보고서*.xlsx"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def acquire_fixed_xlsx(source: Path | None = None) -> tuple[Path, str]:
    """Return (fixed_path, source_desc).

    If `source` is provided, it will be copied into the fixed Downloads path.
    """
    if source is not None:
        source = source.expanduser().resolve()
        if not source.exists():
            raise FlowStop(f"--source 파일이 존재하지 않음: {source}")
        shutil.copy2(source, FIXED_XLSX)
        return FIXED_XLSX, f"copied_from_source:{source}"

    if FIXED_XLSX.exists():
        return FIXED_XLSX, f"fixed:{FIXED_XLSX}"

    src = find_newest_user_download()
    if not src:
        raise FlowStop(
            f"입력 XLSX를 찾을 수 없음. 기대 경로: {FIXED_XLSX} 또는 ~/Downloads/일자별-보고서*.xlsx"
        )
    shutil.copy2(src, FIXED_XLSX)
    return FIXED_XLSX, f"copied_from:{src}"


def extract_values(xlsx_path: Path, expected_date: str) -> ExtractResult:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Raw Data Report"] if "Raw Data Report" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    req = [
        "캠페인 이름",
        "광고 세트 이름",
        "광고 이름",
        "일",
        "노출",
        "통화",
        "지출 금액 (KRW)",
        "클릭(전체)",
        "CPC(전체)",
        "CPM(1,000회 노출당 비용)",
        "CTR(전체)",
        "웹사이트 URL",
    ]
    miss = [h for h in req if h not in idx]
    if miss:
        raise FlowStop("XLSX 컬럼 누락: " + ", ".join(miss))

    rows: list[list] = []
    dates: set[str] = set()

    def get(r, name):
        return r[idx[name]].value

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if all(c.value is None for c in r):
            continue
        date = str(get(r, "일"))[:10]
        dates.add(date)

        ctr = get(r, "CTR(전체)")
        ctr_frac = (float(ctr) / 100.0) if ctr is not None else ""

        rows.append(
            [
                get(r, "캠페인 이름") or "",
                get(r, "광고 세트 이름") or "",
                get(r, "광고 이름") or "",
                date,
                get(r, "노출") or 0,
                get(r, "통화") or "",
                get(r, "지출 금액 (KRW)") or 0,
                get(r, "클릭(전체)") or 0,
                get(r, "CPC(전체)") or 0,
                get(r, "CPM(1,000회 노출당 비용)") or 0,
                ctr_frac,
                get(r, "웹사이트 URL") or "",
                "",  # 모델명 수식
            ]
        )

    if len(dates) != 1:
        raise FlowStop(f"XLSX 날짜가 단일일이 아님: {sorted(dates)}")

    report_date = next(iter(dates))
    if report_date != expected_date:
        raise FlowStop(
            f"날짜 검증 실패: expected={expected_date} but xlsx={report_date} (UI/계정 timezone 영향 가능)."
        )

    return ExtractResult(report_date=report_date, values=rows)


def sheet_last_row() -> int:
    # A column should always be filled for data rows.
    data = gog_get(f"{TAB}!A1:A50000").get("values", [])
    return len(data)


def sheet_has_date(report_date: str) -> bool:
    """Return True if report_date already exists in sheet column D.

    Handles both formatted strings ("YYYY-MM-DD") and numeric date serials.
    """
    data = gog_get(f"{TAB}!D2:D50000").get("values", [])

    for row in data:
        if not row:
            continue
        v = str(row[0]).strip()
        if not v:
            continue

        # Common: formatted string
        if v == report_date or v.startswith(report_date):
            return True

        # Sometimes Sheets APIs may return numeric serials for dates
        try:
            # Google Sheets serial dates are days since 1899-12-30
            if v.replace(".", "", 1).isdigit():
                days = float(v)
                base = datetime(1899, 12, 30)
                d = (base + timedelta(days=days)).date().isoformat()
                if d == report_date:
                    return True
        except Exception:
            pass

    return False


def build_values_with_formula(extracted: ExtractResult, start_row: int) -> list[list]:
    values = []
    for i, row in enumerate(extracted.values):
        sheet_row = start_row + i
        row[12] = f"=IFERROR(VLOOKUP(L{sheet_row},'색인_제품'!A:B,2,FALSE),\"\")"
        values.append(row)
    return values


def send_email(report_date: str) -> str:
    sheet_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit"
    subject = f"[메타광고 일자별 보고서] {report_date} 데이터 업데이트 완료 (Google Sheets 링크)"
    body = (
        "안녕하세요.\n"
        f"메타 광고 일자별 보고서({report_date}, KST) 기준으로 데이터 업데이트 완료된 구글 시트 링크 공유드립니다.\n\n"
        f"- 문서 링크: {sheet_url}\n\n"
        "확인 부탁드립니다.\n"
        "감사합니다.\n"
        "semolabsbot\n"
    )

    (OUT_DIR / "email_subject.txt").write_text(subject, encoding="utf-8")
    (OUT_DIR / "email_body.txt").write_text(body, encoding="utf-8")

    out = sh(
        [
            "node",
            "skills/imap-smtp-email/scripts/smtp.js",
            "send",
            "--to",
            "realtiger@wekeepgrowing.com",
            "--subject",
            subject,
            "--body-file",
            str(OUT_DIR / "email_body.txt"),
            "--from",
            "semolabsbot@gmail.com",
        ]
    )
    # smtp.js prints JSON to stdout
    try:
        j = json.loads(out)
        return j.get("messageId", "")
    except Exception:
        return out.strip()[:500]


def cleanup_fixed() -> None:
    try:
        if FIXED_XLSX.exists():
            FIXED_XLSX.unlink()
    except Exception:
        pass


def main() -> int:
    # Default expected date: yesterday(KST). Can be overridden for cron batch runs.
    expected = kst_yesterday()
    if "--date" in sys.argv:
        i = sys.argv.index("--date")
        try:
            expected = sys.argv[i + 1]
        except Exception:
            raise FlowStop("--date YYYY-MM-DD 형식으로 입력 필요")

    (OUT_DIR / "meta_url.txt").write_text(META_URL, encoding="utf-8")
    (OUT_DIR / "expected_date.txt").write_text(expected, encoding="utf-8")

    # Optional: allow a specific source XLSX (useful when export/download step is manual)
    source: Path | None = None
    if "--source" in sys.argv:
        i = sys.argv.index("--source")
        try:
            source = Path(sys.argv[i + 1])
        except Exception:
            raise FlowStop("--source <path> 형식으로 입력 필요")

    src_desc = ""
    try:
        xlsx_path, src_desc = acquire_fixed_xlsx(source)
        (OUT_DIR / "input_source.txt").write_text(src_desc, encoding="utf-8")
        # Keep an audit copy
        shutil.copy2(xlsx_path, OUT_DIR / "meta_export.xlsx")

        extracted = extract_values(xlsx_path, expected)

        if sheet_has_date(extracted.report_date):
            raise FlowStop(f"중복 방지: 시트 D열에 {extracted.report_date}가 이미 존재 → append 중단")

        last = sheet_last_row()
        start_row = last + 1
        values = build_values_with_formula(extracted, start_row)
        end_row = start_row + len(values) - 1
        target_range = f"{TAB}!A{start_row}:M{end_row}"

        (OUT_DIR / "append_range.txt").write_text(target_range, encoding="utf-8")
        (OUT_DIR / "values.json").write_text(json.dumps(values, ensure_ascii=False), encoding="utf-8")

        gog_update(target_range, values, user_entered=True)

        # formats (whole columns)
        gog_format(
            f"{TAB}!G1:J2000",
            {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}},
            "userEnteredFormat.numberFormat",
        )
        gog_format(
            f"{TAB}!K1:K2000",
            {"numberFormat": {"type": "PERCENT", "pattern": "0.0%"}},
            "userEnteredFormat.numberFormat",
        )

        message_id = send_email(extracted.report_date)
        (OUT_DIR / "email_message_id.txt").write_text(str(message_id), encoding="utf-8")

        print(json.dumps({
            "status": "ok",
            "expected": expected,
            "report_date": extracted.report_date,
            "rows": len(values),
            "range": target_range,
            "email_message_id": message_id,
            "input_source": src_desc,
            "out_dir": str(OUT_DIR),
        }, ensure_ascii=False, indent=2))
        return 0

    except FlowStop as e:
        (OUT_DIR / "status.txt").write_text("stopped\n" + str(e) + "\n" + src_desc, encoding="utf-8")
        print(json.dumps({
            "status": "stopped",
            "expected": expected,
            "reason": str(e),
            "input_source": src_desc,
            "out_dir": str(OUT_DIR),
        }, ensure_ascii=False, indent=2))
        return 2

    finally:
        cleanup_fixed()


if __name__ == "__main__":
    raise SystemExit(main())
